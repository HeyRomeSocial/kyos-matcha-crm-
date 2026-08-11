"""
Kyos Matcha — B2B CRM Report Engine
====================================
Parses either the "Download Report" CSV export from the Partner Hub app, or the
richer "Kyos Matcha CRM.xlsx" / live Google Sheet export (Raw Orders + Raw
Partners tabs), and computes weekly/monthly B2B KPIs.

Definitions (locked, matching Ellis's May/June 2026 reports):
  Revenue                = sum of invoice Total for invoices dated within the period
  Order                  = one invoice
  Active account (week)  = a partner with >=1 invoice dated within the reporting week
  Active account (month) = a partner with >=1 invoice dated within the calendar month
                            to date — shown on every weekly report for context, not
                            just the week's own number
  Onboarded/new account  = a partner whose FIRST-EVER invoice falls within the period
                            AND who had no real prior relationship (Hist Orders == 0,
                            no CRM orders before the period start, AND Added to CRM
                            date is within/near the period). If Added to CRM predates
                            the period by more than REACTIVATION_GAP_DAYS, this is a
                            REACTIVATION, not a new onboard — the partner was already
                            known/being worked with, they just hadn't ordered yet.
                            Confirmed real case: Strutt Bakehouse (Derby) and The Salad
                            Lab placed their first CRM invoice in July but were added
                            to the CRM back in early June — Rome confirmed these are
                            returning partners who'd gone quiet, not new contacts.
  Reactivated account     = first invoice this period, but a prior relationship existed
                            (see above) — "welcome back", not "new customer"
  Samples sent (period)  = partners whose Sample Sent Date falls within the period
  No order in 21+ days   = an active partner whose Last Order Date is 21+ days before
                            "today" — a follow-up / win-back candidate, not a churn
                            verdict
  Monthly kg goal         = cumulative kg shipped this calendar month vs a DYNAMIC
                            target = last calendar month's ACTUAL kg shipped + 58kg
                            (Ellis's growth target). Anchor confirmed: June 2026
                            actually ended at 150.5kg, so July's target is ~208.5kg.
                            DEFAULT_TARGET_LADDER is only a fallback for when the
                            prior month's real total isn't known yet — see
                            compute_goal() / compute_full_weekly_report()
  Week-over-week (WoW)    = this week's revenue/orders/kg/active-accounts vs the
                            immediately prior Mon-Sun week — see compute_wow_deltas().
                            A standalone snapshot doesn't say whether things are
                            improving; this gives every headline number a direction.

Usage:
    python3 kyos_report_engine.py <file.xlsx|csv> --start 2026-07-13 --end 2026-07-19 \
        [--today 2026-07-19] [--json out.json]
"""
import csv
import io
import json
import sys
import argparse
from datetime import datetime, date, timedelta
from collections import defaultdict


DEFAULT_TARGET_LADDER = [
    {"month": "July 2026", "target_kg": 208},
    {"month": "August 2026", "target_kg": 266},
    {"month": "September 2026", "target_kg": 324},
    {"month": "October 2026", "target_kg": 382},
    {"month": "November 2026", "target_kg": 440},
    {"month": "December 2026", "target_kg": 498},
    {"month": "January 2027", "target_kg": 556},
]


def parse_export(path):
    """Parse a Partner Hub CSV export into partners[] and invoices[] lists."""
    with open(path, newline='', encoding='utf-8-sig') as f:
        raw = f.read()

    lines = raw.splitlines()
    partners_start = next(i for i, l in enumerate(lines) if l.strip() == 'PARTNERS')
    invoices_start = next(i for i, l in enumerate(lines) if l.strip() == 'INVOICES')

    partners_block = '\n'.join(lines[partners_start + 1: invoices_start]).strip()
    invoices_block = '\n'.join(lines[invoices_start + 1:]).strip()

    partners = list(csv.DictReader(io.StringIO(partners_block)))
    invoices = list(csv.DictReader(io.StringIO(invoices_block)))

    def to_float(v, default=0.0):
        try:
            return float(v)
        except (ValueError, TypeError):
            return default

    def to_int(v, default=0):
        try:
            return int(v)
        except (ValueError, TypeError):
            return default

    for p in partners:
        p['Total Orders'] = to_int(p.get('Total Orders'))
        p['Total KG'] = to_float(p.get('Total KG'))
        p['Total Spent'] = to_float(p.get('Total Spent'))
        p['Price/KG'] = to_float(p.get('Price/KG'), None)

    for inv in invoices:
        inv['Subtotal'] = to_float(inv.get('Subtotal'))
        inv['Total'] = to_float(inv.get('Total'))
        try:
            inv['_date'] = datetime.strptime(inv['Date'].strip(), '%Y-%m-%d').date()
        except Exception:
            inv['_date'] = None

    return {'partners': partners, 'invoices': invoices}


def _norm(name):
    return (name or '').strip().lower()


def compute_period_metrics(data, start, end, prior_data=None, all_invoices_extra=None):
    """CSV-export path (Partner Hub). Kept for fallback / one-off use."""
    invoices = data['invoices']
    partners = data['partners']

    period_invoices = [i for i in invoices if i['_date'] and start <= i['_date'] <= end]

    revenue = sum(i['Total'] for i in period_invoices)
    orders = len(period_invoices)
    paid = sum(i['Total'] for i in period_invoices if i.get('Status') == 'paid')
    unpaid = sum(i['Total'] for i in period_invoices if i.get('Status') == 'unpaid')
    overdue = sum(i['Total'] for i in period_invoices if i.get('Status') == 'overdue')

    active_partner_names = {_norm(i['Partner']) for i in period_invoices}
    active_accounts = len(active_partner_names)

    history = list(invoices)
    if all_invoices_extra:
        seen_keys = {(_norm(i['Partner']), i['_date'], i['Total']) for i in history}
        for i in all_invoices_extra:
            key = (_norm(i['Partner']), i.get('_date'), i.get('Total'))
            if key not in seen_keys:
                history.append(i)
                seen_keys.add(key)

    invoices_by_partner = defaultdict(list)
    for i in history:
        if i['_date']:
            invoices_by_partner[_norm(i['Partner'])].append(i['_date'])

    partner_lookup = {_norm(p['Name']): p for p in partners}

    new_accounts = []
    for name in active_partner_names:
        dates_known = sorted(invoices_by_partner.get(name, []))
        p = partner_lookup.get(name)
        total_orders_alltime = p['Total Orders'] if p else None
        if dates_known:
            earliest = dates_known[0]
            covers_full_history = (total_orders_alltime is not None and len(dates_known) >= total_orders_alltime)
            if start <= earliest <= end and covers_full_history:
                new_accounts.append(name)

    samples_sent_new = None
    current_sample_names = {_norm(p['Name']) for p in partners if p.get('Status') == 'sample_sent'}
    if prior_data is not None:
        prior_names = {_norm(p['Name']) for p in prior_data['partners']}
        samples_sent_new = sorted(n for n in current_sample_names if n not in prior_names)

    total_active_all = sum(1 for p in partners if p.get('Status') == 'active')
    total_sample_sent_all = sum(1 for p in partners if p.get('Status') == 'sample_sent')

    revenue_by_partner = defaultdict(float)
    for i in period_invoices:
        revenue_by_partner[i['Partner']] += i['Total']
    top_partners = sorted(revenue_by_partner.items(), key=lambda x: -x[1])[:5]

    return {
        'period_start': start.isoformat(),
        'period_end': end.isoformat(),
        'revenue': round(revenue, 2),
        'revenue_paid': round(paid, 2),
        'revenue_unpaid': round(unpaid, 2),
        'revenue_overdue': round(overdue, 2),
        'orders': orders,
        'active_accounts': active_accounts,
        'new_accounts_onboarded': sorted(new_accounts),
        'new_accounts_count': len(new_accounts),
        'samples_sent_new': samples_sent_new,
        'samples_sent_new_count': (len(samples_sent_new) if samples_sent_new is not None else None),
        'total_active_accounts_alltime': total_active_all,
        'total_sample_sent_alltime': total_sample_sent_all,
        'aov': round(revenue / orders, 2) if orders else 0,
        'revenue_per_active_account': round(revenue / active_accounts, 2) if active_accounts else 0,
        'top_partners_by_revenue': [(name, round(amt, 2)) for name, amt in top_partners],
    }


def parse_xlsx(path):
    """
    Parse the richer 'Kyos Matcha CRM.xlsx' workbook (Raw Orders + Raw Partners tabs).
    This is the preferred, authoritative source — exact KG per order, per-partner
    historic-vs-CRM order counts, and exact Sample Sent / Added to CRM / Last Order
    dates, so periods can be computed precisely instead of heuristically.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ro = wb['Raw Orders']
    rp = wb['Raw Partners']

    orders = []
    for r in ro.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        orders.append({
            'invoice': r[0], 'partner': (r[1] or '').strip(), 'date': r[2].date() if r[2] else None,
            'status': r[3], 'paid_revenue': r[4] or 0, 'outstanding': r[5] or 0, 'total': r[6] or 0,
            'kg': r[7] or 0,
        })

    partners = []
    for r in rp.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        partners.append({
            'name': r[0].strip(), 'status': r[1], 'hist_orders': r[5] or 0,
            'crm_paid_orders': r[6] or 0, 'crm_unpaid_orders': r[7] or 0, 'total_orders': r[8] or 0,
            'sample_sent_date': r[16].date() if r[16] else None,
            'last_order_date': r[17].date() if r[17] else None,
            'added_to_crm': r[18].date() if r[18] else None,
        })

    return {'orders': orders, 'partners': partners}


REACTIVATION_GAP_DAYS = 14  # if Added to CRM predates the period start by more than this,
                             # a first invoice this period is a REACTIVATION, not a new onboard —
                             # they were already a known/tracked partner before they ordered.


def compute_period_metrics_xlsx(data, start, end):
    """Compute period KPIs from parse_xlsx() output for the [start, end] date range (inclusive)."""
    orders = data['orders']
    partners = {p['name']: p for p in data['partners']}

    period_orders = [o for o in orders if o['date'] and start <= o['date'] <= end]
    revenue = sum(o['total'] for o in period_orders)
    revenue_paid = sum(o['total'] for o in period_orders if o['status'] == 'paid')
    revenue_unpaid = sum(o['total'] for o in period_orders if o['status'] == 'unpaid')
    kg_sold = sum(o['kg'] for o in period_orders)
    orders_count = len(period_orders)

    orders_in_period_by_partner = defaultdict(int)
    revenue_by_partner = defaultdict(float)
    for o in period_orders:
        orders_in_period_by_partner[o['partner']] += 1
        revenue_by_partner[o['partner']] += o['total']

    active_accounts = len(orders_in_period_by_partner)

    # A partner's FIRST invoice this period only counts as a new onboard if there's no
    # prior relationship at all. Hist Orders = 0 alone isn't enough — a partner can have
    # zero paid orders but still have been sampled/worked with weeks earlier (tracked via
    # Added to CRM date). If that date predates the period by more than the gap below,
    # this is a partner coming BACK after a quiet stretch, not a new relationship.
    new_accounts = []
    reactivated_accounts = []
    for name, n_in_period in orders_in_period_by_partner.items():
        p = partners.get(name)
        if not p:
            continue
        orders_before = (p['total_orders'] or 0) - n_in_period
        if (p['hist_orders'] or 0) == 0 and orders_before <= 0:
            added = p.get('added_to_crm')
            if added and (start - added).days > REACTIVATION_GAP_DAYS:
                reactivated_accounts.append(name)
            else:
                new_accounts.append(name)

    samples_sent = [p['name'] for p in data['partners'] if p['sample_sent_date'] and start <= p['sample_sent_date'] <= end]

    top_partners = sorted(revenue_by_partner.items(), key=lambda x: -x[1])[:5]

    return {
        'period_start': start.isoformat(), 'period_end': end.isoformat(),
        'revenue': round(revenue, 2), 'revenue_paid': round(revenue_paid, 2),
        'revenue_unpaid': round(revenue_unpaid, 2), 'revenue_overdue': 0,
        'kg_sold': round(kg_sold, 2),
        'orders': orders_count, 'active_accounts': active_accounts,
        'new_accounts_onboarded': sorted(new_accounts), 'new_accounts_count': len(new_accounts),
        'reactivated_onboarded': sorted(reactivated_accounts), 'reactivated_count': len(reactivated_accounts),
        'samples_sent_new': sorted(samples_sent), 'samples_sent_new_count': len(samples_sent),
        'total_active_accounts_alltime': sum(1 for p in data['partners'] if p['status'] == 'active'),
        'total_sample_sent_alltime': sum(1 for p in data['partners'] if p['status'] == 'sample_sent'),
        'aov': round(revenue / orders_count, 2) if orders_count else 0,
        'revenue_per_active_account': round(revenue / active_accounts, 2) if active_accounts else 0,
        'top_partners_by_revenue': [(n, round(a, 2)) for n, a in top_partners],
    }


def _week_bounds(d):
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def compute_trend_weeks(orders, end_date, n_weeks=8):
    """Kg + revenue for the n_weeks ending with the week containing end_date."""
    cur_start, cur_end = _week_bounds(end_date)
    weeks = []
    for i in range(n_weeks - 1, -1, -1):
        w_start = cur_start - timedelta(weeks=i)
        w_end = cur_end - timedelta(weeks=i)
        wk_orders = [o for o in orders if o['date'] and w_start <= o['date'] <= w_end]
        kg = sum(o['kg'] for o in wk_orders)
        rev = sum(o['total'] for o in wk_orders)
        iso_week = w_start.isocalendar()[1]
        weeks.append({
            'label': f"W{iso_week}", 'start': w_start.isoformat(), 'end': w_end.isoformat(),
            'kg': round(kg, 2), 'revenue': round(rev, 2),
        })
    return weeks


def compute_at_risk(partners, today, min_days=21, top_n=8):
    """Active partners whose last order was min_days+ ago — follow-up candidates, not a churn verdict."""
    rows = []
    for p in partners:
        if p.get('status') != 'active':
            continue
        lod = p.get('last_order_date')
        if not lod:
            continue
        days = (today - lod).days
        if days >= min_days:
            rows.append({'name': p['name'], 'days': days, 'last_order_date': lod.isoformat()})
    rows.sort(key=lambda r: -r['days'])
    return rows[:top_n]


def compute_month_metrics(orders, partners, month_start, month_end_capped):
    month_orders = [o for o in orders if o['date'] and month_start <= o['date'] <= month_end_capped]
    revenue_by_partner = defaultdict(float)
    orders_by_partner = defaultdict(int)
    kg_total = 0.0
    for o in month_orders:
        revenue_by_partner[o['partner']] += o['total']
        orders_by_partner[o['partner']] += 1
        kg_total += o['kg']
    top5 = sorted(revenue_by_partner.items(), key=lambda x: -x[1])[:5]
    top_partners_this_month = [
        {'name': name, 'orders': orders_by_partner[name], 'revenue': round(amt, 2)}
        for name, amt in top5
    ]
    return {
        'active_accounts_month': len(revenue_by_partner),
        'kg_sold_month': round(kg_total, 2),
        'top_partners_this_month': top_partners_this_month,
    }


def compute_goal(mtd_kg, month_label, target_ladder, prior_month_kg=None, prior_month_label=None, increment_kg=58):
    """
    The real growth target is "+58kg over whatever we actually shipped last month",
    not a fixed number typed into a table in advance. When prior_month_kg is known
    (the ACTUAL total from the prior calendar month, cross-checked against Raw:
    Orders — not blindly trusted from a pre-aggregated sheet column), the target is
    computed dynamically as prior_month_kg + increment_kg. DEFAULT_TARGET_LADDER is
    only a fallback for when the prior month's real total isn't available yet.
    Confirmed anchor: June 2026 actually ended at 150.5kg, which is why July's
    target lands at ~208.5kg — matching the ladder by design, not coincidence.
    """
    dynamic_target = round(prior_month_kg + increment_kg, 1) if prior_month_kg is not None else None
    static_target = next((t['target_kg'] for t in target_ladder if t['month'] == month_label), None)
    target = dynamic_target if dynamic_target is not None else static_target
    if target is None:
        return None
    pct = round(min(999, (mtd_kg / target) * 100), 1) if target else 0
    remaining = round(max(0, target - mtd_kg), 2)
    # A stretch marker, not a second hard target — for framing "how far past the
    # minimum could we push" rather than treating +58kg as a ceiling.
    stretch_kg = round(target * 1.15, 1)
    # Hitting the floor mid-month is common once the pace is established — the deck
    # should say so plainly (goal_reached) and immediately reframe toward the next
    # milestone (the stretch goal) rather than just sitting at "100%+".
    goal_reached = mtd_kg >= target
    remaining_to_stretch_kg = round(max(0, stretch_kg - mtd_kg), 2)
    return {
        'month_label': month_label, 'target_kg': target, 'mtd_kg': round(mtd_kg, 2),
        'pct': pct, 'remaining_kg': remaining,
        'prior_month_kg': prior_month_kg, 'prior_month_label': prior_month_label,
        'increment_kg': increment_kg, 'stretch_kg': stretch_kg,
        'goal_reached': goal_reached, 'remaining_to_stretch_kg': remaining_to_stretch_kg,
    }


def _month_end(month_start):
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    return next_month - timedelta(days=1)


def _parse_month_label(label):
    return datetime.strptime(label, '%B %Y')


def _month_diff(a_label, b_label):
    """Months from b_label to a_label (positive if a_label is later)."""
    da, db = _parse_month_label(a_label), _parse_month_label(b_label)
    return (da.year - db.year) * 12 + (da.month - db.month)


def annotate_ladder_with_actuals(target_ladder, monthly_actuals=None, current_month_label=None,
                                   current_mtd_kg=None, increment_kg=58):
    """
    Turns the flat target ladder into a growth trajectory: past months show their
    REAL shipped total (actual_kg), the current month shows progress-so-far
    (actual_kg = mtd_kg, in_progress = True), and future months are target-only.
    monthly_actuals: {"June 2026": 150.5, ...} — verified totals, not guesses.

    DEFAULT_TARGET_LADDER (and any custom ladder passed in) usually starts at the
    CURRENT month, with nothing for months before it — so a verified past actual
    (e.g. June's 150.5kg, passed via monthly_actuals) would otherwise have nowhere
    to attach. To make "kg keeps increasing period over period" visible, any
    monthly_actuals month that falls BEFORE the ladder's first rung gets prepended,
    with its target back-computed by subtracting increment_kg per month from the
    first rung (linear extrapolation of the same +58kg/month climb, backward).
    """
    monthly_actuals = monthly_actuals or {}
    ladder = [dict(t) for t in target_ladder]

    if ladder:
        first_month, first_target = ladder[0]['month'], ladder[0]['target_kg']
        existing_months = {t['month'] for t in ladder}
        missing_before = [m for m in monthly_actuals if m not in existing_months and _month_diff(first_month, m) > 0]
        missing_before.sort(key=lambda m: _month_diff(first_month, m))  # closest to the ladder first
        for m in missing_before:
            dist = _month_diff(first_month, m)
            target_kg = round(first_target - increment_kg * dist, 1)
            ladder.insert(0, {'month': m, 'target_kg': target_kg})

    out = []
    for t in ladder:
        row = dict(t)
        if t['month'] == current_month_label and current_mtd_kg is not None:
            row['actual_kg'] = current_mtd_kg
            row['in_progress'] = True
        elif t['month'] in monthly_actuals:
            row['actual_kg'] = monthly_actuals[t['month']]
            row['in_progress'] = False
        out.append(row)
    return out


def build_vision_summary(target_ladder, current_month_label):
    """One line projecting the trajectory to the ladder's final month, e.g.
    'At this pace, monthly shipments reach 556.5 kg by January 2027 — 3.7x June's total.'
    Needs at least one real actual_kg entry to anchor the multiple; otherwise a plainer
    sentence without the multiplier."""
    if not target_ladder:
        return None
    last = target_ladder[-1]
    anchor = next((t for t in target_ladder if t.get('actual_kg') and not t.get('in_progress')), None)
    idx_current = next((i for i, t in enumerate(target_ladder) if t['month'] == current_month_label), 0)
    months_ahead = max(0, len(target_ladder) - 1 - idx_current)
    base = f"At the +58kg/month pace, monthly shipments reach {last['target_kg']:.1f} kg by {last['month']}"
    if months_ahead:
        base += f" — about {months_ahead} months from now"
    if anchor and anchor['actual_kg']:
        multiple = last['target_kg'] / anchor['actual_kg']
        base += f", roughly {multiple:.1f}x {anchor['month']}'s {anchor['actual_kg']:.1f} kg"
    return base + "."


def compute_wow_deltas(data, week_start, week_end, current):
    """
    Week-over-week context: how does this week's revenue/orders/kg/active-accounts
    compare to the immediately prior Mon-Sun week? A standalone snapshot doesn't say
    whether things are improving — this answers that in one pass, reusing
    compute_period_metrics_xlsx() on the prior 7-day window. `current` is this
    week's already-computed base metrics dict (avoids recomputing it).
    """
    prior_start = week_start - timedelta(days=7)
    prior_end = week_end - timedelta(days=7)
    prior = compute_period_metrics_xlsx(data, prior_start, prior_end)

    def pct(cur, prev):
        if not prev:
            return None
        return round((cur - prev) / prev * 100, 1)

    return {
        'revenue_prior': prior['revenue'], 'revenue_pct': pct(current['revenue'], prior['revenue']),
        'orders_prior': prior['orders'], 'orders_pct': pct(current['orders'], prior['orders']),
        'kg_prior': prior['kg_sold'], 'kg_pct': pct(current['kg_sold'], prior['kg_sold']),
        'active_prior': prior['active_accounts'], 'active_pct': pct(current['active_accounts'], prior['active_accounts']),
    }


def compute_full_weekly_report(data, week_start, week_end, today=None, target_ladder=None,
                                 prior_month_kg=None, prior_month_label=None, monthly_actuals=None):
    """
    The full weekly report shape: weekly KPIs + month-to-date active cafés, top
    cafés this month, a monthly kg growth trajectory (actual-vs-target ladder, no
    chart — see annotate_ladder_with_actuals/build_vision_summary), 21+ day
    follow-up list, and monthly kg goal progress. Pass prior_month_kg (the ACTUAL
    total shipped last calendar month, verified against Raw: Orders) so the goal
    is anchored to real performance (target = prior_month_kg + 58) rather than a
    static pre-typed table — see compute_goal(). Pass monthly_actuals for a
    {month_label: actual_kg} map covering any other verified past months, so the
    trajectory shows real growth, not just future targets. Falls back to
    DEFAULT_TARGET_LADDER if prior_month_kg is None.
    Insights are NOT computed here — they require judgment (operational context,
    what to prioritize) and should be authored on top of this data, not derived
    mechanically.
    """
    if today is None:
        today = week_end
    if target_ladder is None:
        target_ladder = DEFAULT_TARGET_LADDER

    base = compute_period_metrics_xlsx(data, week_start, week_end)

    month_start = week_end.replace(day=1)
    month_end = _month_end(month_start)
    month_end_capped = min(month_end, today)
    month_metrics = compute_month_metrics(data['orders'], data['partners'], month_start, month_end_capped)

    trend = compute_trend_weeks(data['orders'], week_end, n_weeks=8)
    at_risk = compute_at_risk(data['partners'], today)
    wow = compute_wow_deltas(data, week_start, week_end, base)
    month_label = month_start.strftime('%B %Y')
    goal = compute_goal(month_metrics['kg_sold_month'], month_label, target_ladder,
                         prior_month_kg=prior_month_kg, prior_month_label=prior_month_label)

    monthly_actuals = dict(monthly_actuals or {})
    if prior_month_kg is not None and prior_month_label:
        monthly_actuals.setdefault(prior_month_label, prior_month_kg)
    trajectory = annotate_ladder_with_actuals(
        target_ladder, monthly_actuals=monthly_actuals,
        current_month_label=month_label, current_mtd_kg=month_metrics['kg_sold_month'],
    )
    vision_summary = build_vision_summary(trajectory, month_label)

    result = dict(base)
    result.update({
        'month_label': month_label,
        'active_accounts_week': base['active_accounts'],
        'active_accounts_month': month_metrics['active_accounts_month'],
        'top_partners_this_month': month_metrics['top_partners_this_month'],
        'trend_8week': trend,
        'at_risk': at_risk,
        'wow': wow,
        'goal': goal,
        'target_ladder': trajectory,
        'vision_summary': vision_summary,
    })
    return result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('data_path', help='CSV export (Partner Hub) or .xlsx (Kyos Matcha CRM workbook)')
    ap.add_argument('--start', required=True)
    ap.add_argument('--end', required=True)
    ap.add_argument('--today', default=None, help='Reference date for at-risk / MTD calc; defaults to --end')
    ap.add_argument('--prior-csv', default=None, help='Only used for CSV mode')
    ap.add_argument('--basic', action='store_true', help='xlsx mode: skip the full report, just weekly KPIs')
    ap.add_argument('--prior-month-kg', type=float, default=None, help='Actual kg shipped last calendar month, for the dynamic +58kg goal')
    ap.add_argument('--prior-month-label', default=None, help='e.g. "June 2026" — label for --prior-month-kg')
    ap.add_argument('--json', default=None)
    args = ap.parse_args()

    start = datetime.strptime(args.start, '%Y-%m-%d').date()
    end = datetime.strptime(args.end, '%Y-%m-%d').date()
    today = datetime.strptime(args.today, '%Y-%m-%d').date() if args.today else end

    if args.data_path.lower().endswith('.xlsx'):
        data = parse_xlsx(args.data_path)
        if args.basic:
            metrics = compute_period_metrics_xlsx(data, start, end)
        else:
            metrics = compute_full_weekly_report(data, start, end, today=today,
                                                   prior_month_kg=args.prior_month_kg,
                                                   prior_month_label=args.prior_month_label)
    else:
        data = parse_export(args.data_path)
        prior_data = parse_export(args.prior_csv) if args.prior_csv else None
        metrics = compute_period_metrics(data, start, end, prior_data=prior_data)

    out = json.dumps(metrics, indent=2)
    print(out)
    if args.json:
        with open(args.json, 'w') as f:
            f.write(out)


if __name__ == '__main__':
    main()
